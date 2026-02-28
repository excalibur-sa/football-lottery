import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import config


# 样式常量
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
DATA_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
DATA_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
ALT_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
SECTION_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

HAFU_LABELS = {
    "win_win": "胜-胜", "win_draw": "胜-平", "win_lose": "胜-负",
    "draw_win": "平-胜", "draw_draw": "平-平", "draw_lose": "平-负",
    "lose_win": "负-胜", "lose_draw": "负-平", "lose_lose": "负-负",
}


def _apply_header_style(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _apply_data_style(ws, row, col_start, col_end, is_alt=False):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = THIN_BORDER
        if is_alt:
            cell.fill = ALT_ROW_FILL


def _auto_column_width(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                # 中文字符按2个宽度计算
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_length = max(max_length, length)
        ws.column_dimensions[col_letter].width = min(max_length + 4, 30)


def _write_summary_sheet(ws, matches):
    """写入比赛汇总Sheet"""
    ws.title = "比赛汇总"
    headers = [
        "序号", "比赛时间", "联赛", "主队", "客队",
        "胜", "平", "负", "让球", "让胜", "让平", "让负"
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, 1, len(headers))
    ws.freeze_panes = "A2"

    for idx, m in enumerate(matches, 1):
        row = idx + 1
        had = m.get("had_odds", {})
        hhad = m.get("hhad_odds", {})
        values = [
            idx,
            m.get("match_time", ""),
            m.get("league", ""),
            m.get("home_team", ""),
            m.get("away_team", ""),
            had.get("win", ""),
            had.get("draw", ""),
            had.get("lose", ""),
            hhad.get("handicap", ""),
            hhad.get("win", ""),
            hhad.get("draw", ""),
            hhad.get("lose", ""),
        ]
        for col, v in enumerate(values, 1):
            ws.cell(row=row, column=col, value=v)
        _apply_data_style(ws, row, 1, len(headers), is_alt=(idx % 2 == 0))

    _auto_column_width(ws)


def _write_detail_sheet(wb, match, index):
    """为单场比赛创建详情Sheet"""
    title = f"{index}-{match.get('home_team', '')}vs{match.get('away_team', '')}"
    # Sheet名最长31字符
    if len(title) > 31:
        title = title[:31]
    ws = wb.create_sheet(title=title)

    row = 1

    # === 基本信息 ===
    ws.cell(row=row, column=1, value="基本信息").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    info_items = [
        ("比赛ID", match.get("match_id", "")),
        ("比赛时间", match.get("match_time", "")),
        ("联赛", match.get("league", "")),
        ("主队", match.get("home_team", "")),
        ("客队", match.get("away_team", "")),
    ]
    for label, val in info_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=val).border = THIN_BORDER
        row += 1

    row += 1

    # === 胜平负赔率 ===
    ws.cell(row=row, column=1, value="胜平负赔率").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1
    had = match.get("had_odds", {})
    for col, h in enumerate(["胜", "平", "负"], 1):
        ws.cell(row=row, column=col, value=h)
    _apply_header_style(ws, row, 1, 3)
    row += 1
    for col, key in enumerate(["win", "draw", "lose"], 1):
        ws.cell(row=row, column=col, value=had.get(key, ""))
    _apply_data_style(ws, row, 1, 3)
    row += 2

    # === 让球胜平负 ===
    ws.cell(row=row, column=1, value="让球胜平负").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    hhad = match.get("hhad_odds", {})
    for col, h in enumerate(["让球数", "胜", "平", "负"], 1):
        ws.cell(row=row, column=col, value=h)
    _apply_header_style(ws, row, 1, 4)
    row += 1
    for col, key in enumerate(["handicap", "win", "draw", "lose"], 1):
        ws.cell(row=row, column=col, value=hhad.get(key, ""))
    _apply_data_style(ws, row, 1, 4)
    row += 2

    # === 比分赔率 ===
    crs = match.get("crs_odds", {})
    if crs:
        ws.cell(row=row, column=1, value="比分赔率").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

        # 分主胜、平、客胜三组
        home_win = {k: v for k, v in crs.items()
                    if len(k.split(":")) == 2 and int(k.split(":")[0]) > int(k.split(":")[1])}
        draws = {k: v for k, v in crs.items()
                 if len(k.split(":")) == 2 and int(k.split(":")[0]) == int(k.split(":")[1])}
        away_win = {k: v for k, v in crs.items()
                    if len(k.split(":")) == 2 and int(k.split(":")[0]) < int(k.split(":")[1])}

        for group_name, group_data in [("主胜", home_win), ("平局", draws), ("客胜", away_win)]:
            if not group_data:
                continue
            ws.cell(row=row, column=1, value=group_name).font = Font(bold=True, italic=True)
            row += 1
            for col, h in enumerate(["比分", "赔率"], 1):
                ws.cell(row=row, column=col, value=h)
            _apply_header_style(ws, row, 1, 2)
            row += 1
            for i, (score, odds) in enumerate(sorted(group_data.items())):
                ws.cell(row=row, column=1, value=score)
                ws.cell(row=row, column=2, value=odds)
                _apply_data_style(ws, row, 1, 2, is_alt=(i % 2 == 0))
                row += 1
            row += 1

    # === 总进球赔率 ===
    ttg = match.get("ttg_odds", {})
    if ttg:
        ws.cell(row=row, column=1, value="总进球赔率").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        for col, h in enumerate(["总进球数", "赔率"], 1):
            ws.cell(row=row, column=col, value=h)
        _apply_header_style(ws, row, 1, 2)
        row += 1
        for i, (goals, odds) in enumerate(sorted(ttg.items(), key=lambda x: x[0])):
            label = f"{goals}球" if goals != "7+" else "7+球"
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=odds)
            _apply_data_style(ws, row, 1, 2, is_alt=(i % 2 == 0))
            row += 1
        row += 1

    # === 半全场赔率 ===
    hafu = match.get("hafu_odds", {})
    if hafu:
        ws.cell(row=row, column=1, value="半全场赔率").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        for col, h in enumerate(["半场-全场", "赔率"], 1):
            ws.cell(row=row, column=col, value=h)
        _apply_header_style(ws, row, 1, 2)
        row += 1
        for i, (key, odds) in enumerate(hafu.items()):
            label = HAFU_LABELS.get(key, key)
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=odds)
            _apply_data_style(ws, row, 1, 2, is_alt=(i % 2 == 0))
            row += 1
        row += 1

    # === 胜平负赔率变化历史 ===
    had_history = match.get("had_history", [])
    if had_history:
        ws.cell(row=row, column=1, value=f"胜平负赔率变化历史（共{len(had_history)}条）").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
        for col, h in enumerate(["更新日期", "更新时间", "胜", "平", "负"], 1):
            ws.cell(row=row, column=col, value=h)
        _apply_header_style(ws, row, 1, 5)
        row += 1
        for i, item in enumerate(had_history):
            ws.cell(row=row, column=1, value=item.get("update_date", ""))
            ws.cell(row=row, column=2, value=item.get("update_time", ""))
            ws.cell(row=row, column=3, value=item.get("win", ""))
            ws.cell(row=row, column=4, value=item.get("draw", ""))
            ws.cell(row=row, column=5, value=item.get("lose", ""))
            _apply_data_style(ws, row, 1, 5, is_alt=(i % 2 == 0))
            row += 1
        row += 1

    # === 让球胜平负赔率变化历史 ===
    hhad_history = match.get("hhad_history", [])
    if hhad_history:
        first_handicap = hhad_history[0].get("handicap", "") if hhad_history else ""
        ws.cell(row=row, column=1, value=f"让球胜平负赔率变化历史（让{first_handicap}球，共{len(hhad_history)}条）").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
        for col, h in enumerate(["更新日期", "更新时间", "让球", "让胜", "让平", "让负"], 1):
            ws.cell(row=row, column=col, value=h)
        _apply_header_style(ws, row, 1, 6)
        row += 1
        for i, item in enumerate(hhad_history):
            ws.cell(row=row, column=1, value=item.get("update_date", ""))
            ws.cell(row=row, column=2, value=item.get("update_time", ""))
            ws.cell(row=row, column=3, value=item.get("handicap", ""))
            ws.cell(row=row, column=4, value=item.get("win", ""))
            ws.cell(row=row, column=5, value=item.get("draw", ""))
            ws.cell(row=row, column=6, value=item.get("lose", ""))
            _apply_data_style(ws, row, 1, 6, is_alt=(i % 2 == 0))
            row += 1
        row += 1

    # === 赔率差值分析 ===
    had_history = match.get("had_history", [])
    hhad_history = match.get("hhad_history", [])
    
    # 判断是否有胜平负指数更新：HAD历史非空即有更新
    has_had_update = len(had_history) >= 1
    
    # 需要有让球更新(>=2条)才能计算赔率差
    if len(hhad_history) >= 2 and has_had_update:
        ws.cell(row=row, column=1, value="赔率差值分析").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
        
        # 表头
        headers = ["胜的赔率的差", "负的赔率的差", "双平赔率的差", "胜差正负", "负差正负", "双平差正负"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h)
        _apply_header_style(ws, row, 1, 6)
        row += 1
        
        # 获取第一条数据作为基准（用于胜差、负差）
        base_win = had_history[0].get("win", 0) or 0
        base_lose = had_history[0].get("lose", 0) or 0
        
        # 构建HAD时间戳索引，用于时间匹配
        def parse_timestamp(item):
            return item.get("update_date", "") + " " + item.get("update_time", "")
        
        def parse_full_datetime(item):
            """将HAD/HHAD记录的日期+时间解析为datetime对象"""
            try:
                from datetime import datetime as dt_cls
                date_str = item.get("update_date", "")
                time_str = item.get("update_time", "")
                return dt_cls.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
            except:
                return None
        
        def time_diff_seconds(t1, t2):
            """计算两个时间字符串的秒数差（仅比较时间部分HH:MM:SS）"""
            try:
                from datetime import datetime
                fmt = "%H:%M:%S"
                dt1 = datetime.strptime(t1.split()[-1] if " " in t1 else t1, fmt)
                dt2 = datetime.strptime(t2.split()[-1] if " " in t2 else t2, fmt)
                return abs((dt1 - dt2).total_seconds())
            except:
                return float('inf')
        
        def find_matching_had(hhad_item, had_list, tolerance=300):
            """在HAD列表中查找与HHAD时间戳最接近的记录（同日期，容差5分钟）"""
            hhad_date = hhad_item.get("update_date", "")
            hhad_time = hhad_item.get("update_time", "")
            best_match = None
            best_diff = float('inf')
            for idx, had in enumerate(had_list):
                if had.get("update_date", "") != hhad_date:
                    continue
                diff = time_diff_seconds(had.get("update_time", ""), hhad_time)
                if diff < best_diff and diff <= tolerance:
                    best_diff = diff
                    best_match = idx
            return best_match
        
        # 对齐规则：
        # 第1行(r=0): 胜/负差取 HAD[1]，双平差取 HAD[0]/HHAD[0]（保持特殊对齐）
        # 后续行: 按时间戳匹配HAD和HHAD，无匹配时使用最近的HAD平数据
        last_had_draw = had_history[-1].get("draw", 0) or 0
        used_had_for_wl = set()  # 记录已用于胜/负差的HAD索引
        had_dd_covered = set()   # 记录已在HHAD循环中计算了双平差的HAD索引
        last_used_had_draw = had_history[0].get("draw", 0) or 0  # 记录最近使用的HAD平数据
        
        total_rows = len(hhad_history)
        
        for r in range(total_rows):
            if r == 0:
                # 第1行特殊对齐：胜/负差取 HAD[1]，双平差取 HAD[0]/HHAD[0]
                if len(had_history) >= 2:
                    current_win = had_history[1].get("win", 0) or 0
                    current_lose = had_history[1].get("lose", 0) or 0
                    win_diff = round(current_win - base_win, 2)
                    lose_diff = round(current_lose - base_lose, 2)
                    ws.cell(row=row, column=1, value=win_diff)
                    ws.cell(row=row, column=2, value=lose_diff)
                    ws.cell(row=row, column=4, value="正" if win_diff >= 0 else "负")
                    ws.cell(row=row, column=5, value="正" if lose_diff >= 0 else "负")
                    used_had_for_wl.add(1)
                else:
                    ws.cell(row=row, column=1, value="-")
                    ws.cell(row=row, column=2, value="-")
                    ws.cell(row=row, column=4, value="-")
                    ws.cell(row=row, column=5, value="-")
                
                cur_had_draw = had_history[0].get("draw", 0) or 0
                cur_hhad_draw = hhad_history[0].get("draw", 0) or 0
                had_dd_covered.add(0)  # HAD[0]的draw已用于双平差
                last_used_had_draw = cur_had_draw
            else:
                # 后续行：按时间戳匹配，匹配不到则使用向前查找最近的HAD
                matched_had_idx = find_matching_had(hhad_history[r], had_history)
                
                if matched_had_idx is not None and matched_had_idx not in used_had_for_wl:
                    # 找到匹配且未使用过
                    current_win = had_history[matched_had_idx].get("win", 0) or 0
                    current_lose = had_history[matched_had_idx].get("lose", 0) or 0
                    win_diff = round(current_win - base_win, 2)
                    lose_diff = round(current_lose - base_lose, 2)
                    ws.cell(row=row, column=1, value=win_diff)
                    ws.cell(row=row, column=2, value=lose_diff)
                    ws.cell(row=row, column=4, value="正" if win_diff >= 0 else "负")
                    ws.cell(row=row, column=5, value="正" if lose_diff >= 0 else "负")
                    used_had_for_wl.add(matched_had_idx)
                    cur_had_draw = had_history[matched_had_idx].get("draw", 0) or 0
                    had_dd_covered.add(matched_had_idx)
                    last_used_had_draw = cur_had_draw
                else:
                    # 无匹配或已使用，胜/负差显示"-"
                    ws.cell(row=row, column=1, value="-")
                    ws.cell(row=row, column=2, value="-")
                    ws.cell(row=row, column=4, value="-")
                    ws.cell(row=row, column=5, value="-")
                    # 双平差：向前查找时间戳<=当前HHAD的最近HAD
                    hhad_dt = parse_full_datetime(hhad_history[r])
                    cur_had_draw = last_used_had_draw  # 默认使用最近一次的
                    if hhad_dt:
                        best_candidate = None
                        best_dt = None
                        for had_item in had_history:
                            had_dt = parse_full_datetime(had_item)
                            if had_dt and had_dt <= hhad_dt:
                                if best_dt is None or had_dt > best_dt:
                                    best_dt = had_dt
                                    best_candidate = had_item
                        if best_candidate:
                            cur_had_draw = best_candidate.get("draw", 0) or 0
                            last_used_had_draw = cur_had_draw
                
                cur_hhad_draw = hhad_history[r].get("draw", 0) or 0
            
            double_draw_diff = round(cur_hhad_draw - cur_had_draw, 2) if cur_hhad_draw and cur_had_draw else 0
            ws.cell(row=row, column=3, value=double_draw_diff)
            ws.cell(row=row, column=6, value="正" if double_draw_diff >= 0 else "负")
            
            _apply_data_style(ws, row, 1, 6, is_alt=(r % 2 == 0))
            row += 1
        
        # 补充未完整覆盖的HAD记录
        row_idx = len(hhad_history)
        for had_idx in range(1, len(had_history)):
            # 已在HHAD循环中同时覆盖了胜/负差和双平差的，跳过
            if had_idx in used_had_for_wl and had_idx in had_dd_covered:
                continue
            
            if had_idx in used_had_for_wl:
                # 胜/负差已在HHAD循环显示，仅需双平差独占一行
                ws.cell(row=row, column=1, value="-")
                ws.cell(row=row, column=2, value="-")
                ws.cell(row=row, column=4, value="-")
                ws.cell(row=row, column=5, value="-")
            else:
                # 完整行：胜/负差 + 双平差
                current_win = had_history[had_idx].get("win", 0) or 0
                current_lose = had_history[had_idx].get("lose", 0) or 0
                win_diff = round(current_win - base_win, 2)
                lose_diff = round(current_lose - base_lose, 2)
                ws.cell(row=row, column=1, value=win_diff)
                ws.cell(row=row, column=2, value=lose_diff)
                ws.cell(row=row, column=4, value="正" if win_diff >= 0 else "负")
                ws.cell(row=row, column=5, value="正" if lose_diff >= 0 else "负")
            
            # 双平差：向前查找最近的HHAD让平赔率
            had_dt = parse_full_datetime(had_history[had_idx])
            nearest_hhad_draw = None
            if had_dt:
                best_candidate = None
                best_dt = None
                for hhad_item in hhad_history:
                    hhad_dt = parse_full_datetime(hhad_item)
                    if hhad_dt and hhad_dt <= had_dt:
                        if best_dt is None or hhad_dt > best_dt:
                            best_dt = hhad_dt
                            best_candidate = hhad_item
                if best_candidate:
                    nearest_hhad_draw = best_candidate.get("draw", 0) or 0
            
            if nearest_hhad_draw is not None:
                cur_had_draw = had_history[had_idx].get("draw", 0) or 0
                double_draw_diff = round(nearest_hhad_draw - cur_had_draw, 2)
                ws.cell(row=row, column=3, value=double_draw_diff)
                ws.cell(row=row, column=6, value="正" if double_draw_diff >= 0 else "负")
            else:
                ws.cell(row=row, column=3, value="-")
                ws.cell(row=row, column=6, value="-")
            
            _apply_data_style(ws, row, 1, 6, is_alt=(row_idx % 2 == 0))
            row += 1
            row_idx += 1

    elif len(hhad_history) >= 2:
        # 无胜平负指数更新，仅用HHAD让平数据做差：HHAD[i].draw - HHAD[0].draw
        ws.cell(row=row, column=1, value="赔率差值分析").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
        
        headers = ["胜的赔率的差", "负的赔率的差", "双平赔率的差", "胜差正负", "负差正负", "双平差正负"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h)
        _apply_header_style(ws, row, 1, 6)
        row += 1
        
        base_hhad_draw = hhad_history[0].get("draw", 0) or 0
        
        for i in range(1, len(hhad_history)):
            ws.cell(row=row, column=1, value="-")
            ws.cell(row=row, column=2, value="-")
            ws.cell(row=row, column=4, value="-")
            ws.cell(row=row, column=5, value="-")
            
            cur_hhad_draw = hhad_history[i].get("draw", 0) or 0
            draw_diff = round(cur_hhad_draw - base_hhad_draw, 2)
            ws.cell(row=row, column=3, value=draw_diff)
            ws.cell(row=row, column=6, value="正" if draw_diff >= 0 else "负")
            
            _apply_data_style(ws, row, 1, 6, is_alt=(i % 2 == 0))
            row += 1

    _auto_column_width(ws)


def generate_excel(matches):
    """生成竞彩足球数据Excel文件

    Args:
        matches: 包含完整赔率的比赛数据列表

    Returns:
        str: 生成的Excel文件完整路径
    """
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)

    wb = Workbook()
    ws_summary = wb.active
    _write_summary_sheet(ws_summary, matches)

    for idx, match in enumerate(matches, 1):
        _write_detail_sheet(wb, match, idx)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"竞彩足球_{timestamp}.xlsx"
    filepath = os.path.join(config.OUTPUT_DIR, filename)
    wb.save(filepath)
    return filepath, filename
